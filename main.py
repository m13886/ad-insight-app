"""
AdInsight AI - AI Smart Report Generator for Freelancers & Agencies
الإصدار 2.0 - مع تكامل OpenAI API
"""

import streamlit as st
import pandas as pd
import numpy as np
from io import BytesIO
import datetime
import os
import time
import random
import uuid
import base64
import tempfile
import re
import json
import secrets
import hashlib
import hmac
import ctypes
import sys
import ast
from pathlib import Path
from openai import OpenAI, APIError, APIConnectionError, RateLimitError, AuthenticationError
from dotenv import load_dotenv
from cryptography.fernet import Fernet

# ---------------------------- تحميل متغيرات البيئة ----------------------------
load_dotenv()

# ---------------------------- Feature Flag للتطوير ----------------------------
DEV_MODE = os.getenv("DEV_MODE", "1") == "1"

# ---------------------------- إعداد مفتاح التشفير من st.secrets ----------------------------
def get_fernet():
    try:
        secret = st.secrets["FERNET_SECRET"]
    except KeyError:
        raise RuntimeError(
            "❌ لم يتم العثور على FERNET_SECRET في secrets.\n"
            "أضف المفتاح في لوحة تحكم Streamlit (App settings -> Secrets) على شكل:\n"
            "FERNET_SECRET = 'your_base64_key_here'\n"
            "يمكنك توليد مفتاح باستخدام: from cryptography.fernet import Fernet; print(Fernet.generate_key().decode())"
        )
    return Fernet(secret.encode())

fernet = get_fernet()

# ---------------------------- مكتبات إنشاء PDF ----------------------------
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image
from reportlab.lib.units import cm
from reportlab.lib.utils import ImageReader

# ---------------------------- ثوابت الحماية ----------------------------
MAX_ROWS = 50000
MAX_COLS = 100

# ---------------------------- تعريف الثوابت والمرادفات ----------------------------
STANDARD_COLUMNS = {
    'campaign_name': 'اسم الحملة',
    'impressions': 'مرات الظهور',
    'clicks': 'النقرات',
    'spend': 'الإنفاق',
    'conversions': 'التحويلات',
    'revenue': 'الإيرادات'
}

SYNONYMS = {
    'campaign_name': [
        'campaign', 'campaign name', 'campaignname', 'ad campaign',
        'اسم الحملة', 'الحملة', 'campaign_id', 'campaign id'
    ],
    'impressions': [
        'impression', 'imps', 'مرات الظهور', 'ظهور', 'views', 'reach'
    ],
    'clicks': [
        'click', 'clicks', 'النقرات', 'نقرات', 'link clicks'
    ],
    'spend': [
        'spend', 'cost', 'amount spent', 'ad spend', 'الإنفاق',
        'المبلغ المنفق', 'التكلفة', 'spent'
    ],
    'conversions': [
        'conversions', 'conversion', 'results', 'التحويلات',
        'تحويلات', 'purchases', 'signups', 'leads'
    ],
    'revenue': [
        'revenue', 'sales', 'purchase revenue', 'conversion value',
        'الإيرادات', 'المبيعات', 'قيمة التحويل'
    ]
}

REQUIRED_COLUMNS = ['campaign_name', 'impressions', 'clicks', 'spend', 'conversions']

# ---------------------------- دوال مساعدة للجهاز والتاريخ ----------------------------
def get_device_id():
    try:
        mac = uuid.getnode()
        if mac == uuid.getnode() and (mac >> 40) % 2 == 0:
            return hashlib.sha256(str(mac).encode()).hexdigest()
    except:
        pass
    fallback = f"{os.name}_{os.getlogin()}_{Path.home()}"
    return hashlib.sha256(fallback.encode()).hexdigest()

# ---------------------------- تخزين بيانات العميل المحلي (مشفر) ----------------------------
CLIENT_DATA_PATH = Path.home() / ".adinsight" / "client_data.json"

def ensure_data_dir():
    CLIENT_DATA_PATH.parent.mkdir(parents=True, exist_ok=True)

def save_client_data_encrypted(data):
    ensure_data_dir()
    json_bytes = json.dumps(data).encode('utf-8')
    encrypted = fernet.encrypt(json_bytes)
    with open(CLIENT_DATA_PATH, "wb") as f:
        f.write(encrypted)

def load_client_data_encrypted():
    if not CLIENT_DATA_PATH.exists():
        raise FileNotFoundError("❌ ملف الترخيص غير موجود. يرجى تفعيل الترخيص أولاً.")
    with open(CLIENT_DATA_PATH, "rb") as f:
        encrypted = f.read()
    try:
        decrypted = fernet.decrypt(encrypted)
        data = json.loads(decrypted.decode('utf-8'))
        if "device_id" not in data:
            data["device_id"] = None
        if "expiry" not in data:
            data["expiry"] = None
        return data
    except Exception as e:
        raise ValueError(
            f"❌ فشل فك تشفير ملف الترخيص. ربما تم التلاعب بالملف أو استخدام مفتاح خاطئ.\n"
            f"التفاصيل: {e}"
        )

def is_trial_valid(data):
    if "device_id" not in data or data["device_id"] is None:
        data["device_id"] = get_device_id()
        save_client_data_encrypted(data)

    if data["device_id"] != get_device_id():
        return False, "⛔ النسخة التجريبية مرتبطة بجهاز آخر ولا يمكن استخدامها هنا."

    first_use_str = data.get("first_use")
    if not first_use_str:
        return False, "⛔ بيانات النسخة التجريبية غير صالحة (first_use مفقود)."
    try:
        first_use = datetime.datetime.fromisoformat(first_use_str)
    except:
        return False, "⛔ تنسيق تاريخ أول استخدام غير صالح."

    trial_days = data.get("trial_days", 7)
    days_passed = (datetime.datetime.now() - first_use).days
    if days_passed > trial_days:
        return False, "⛔ انتهت صلاحية النسخة التجريبية (تجاوزت المدة المسموحة)."

    usage = data.get("usage_count", 0)
    limit = data.get("trial_limit", 10)
    if usage >= limit:
        return False, f"⛔ انتهت النسخة التجريبية (استخدمت {usage} من أصل {limit} استخدامات)."

    data["usage_count"] = usage + 1
    save_client_data_encrypted(data)

    remaining_usage = limit - (usage + 1)
    remaining_days = trial_days - days_passed
    return True, f"⚠️ نسخة تجريبية: {remaining_usage} استخدامات متبقية، {remaining_days} أيام متبقية"

# ---------------------------- إدارة مفاتيح الترخيص ----------------------------
ALLOWED_KEYS = {
    "A1B2-C3D4-E5F6-G7H8": {"type": "lifetime", "email": "client1@example.com"},
    "I9J0-K1L2-M3N4-O5P6": {"type": "trial", "email": "client2@example.com"},
}

def generate_license_key(length=16):
    alphabet = "ABCDEFGHIJKLMNOPQRSTUVWXYZ0123456789"
    parts = []
    for _ in range(4):
        part = ''.join(secrets.choice(alphabet) for _ in range(4))
        parts.append(part)
    return '-'.join(parts)

def verify_license_key(input_key):
    return input_key in ALLOWED_KEYS

def generate_signature(license_key: str) -> str:
    return hmac.new(fernet._encryption_key, license_key.encode(), hashlib.sha256).hexdigest()

def activate_license(key, expiry_days=365):
    if not verify_license_key(key):
        return False, "❌ مفتاح غير صالح."

    try:
        data = load_client_data_encrypted()
    except FileNotFoundError:
        data = {
            "email": "",
            "license_status": "trial",
            "usage_count": 0,
            "trial_limit": 10,
            "first_use": None,
            "trial_days": 7,
            "license_key": None,
            "license_type": None,
            "signature": None,
            "device_id": None,
            "expiry": None
        }
    except ValueError as e:
        data = {
            "email": "",
            "license_status": "trial",
            "usage_count": 0,
            "trial_limit": 10,
            "first_use": None,
            "trial_days": 7,
            "license_key": None,
            "license_type": None,
            "signature": None,
            "device_id": None,
            "expiry": None
        }

    device_id = get_device_id()
    expiry = (datetime.date.today() + datetime.timedelta(days=expiry_days)).isoformat()

    data["license_status"] = "active"
    data["license_key"] = key
    data["device_id"] = device_id
    data["expiry"] = expiry
    data["license_type"] = ALLOWED_KEYS[key]["type"]
    data["email"] = ALLOWED_KEYS[key].get("email", data.get("email", ""))
    data["signature"] = generate_signature(key)

    save_client_data_encrypted(data)
    return True, "✅ تم التفعيل بنجاح!"

def verify_license_signature() -> bool:
    try:
        data = load_client_data_encrypted()
    except:
        return False
    key = data.get("license_key", "")
    sig = data.get("signature", "")
    if not key or not sig:
        return False
    expected_sig = generate_signature(key)
    return hmac.compare_digest(sig, expected_sig)

def check_license_secure_with_trial():
    if DEV_MODE:
        return True, "⚠️ وضع التطوير: تم تجاوز التحقق من الترخيص."

    try:
        data = load_client_data_encrypted()
    except (FileNotFoundError, ValueError) as e:
        return False, str(e)

    status = data.get("license_status", "trial")

    if status == "active":
        expiry = data.get("expiry")
        device_id = data.get("device_id")
        stored_key = data.get("license_key")

        if not expiry or not device_id or not stored_key:
            return False, "⛔ بيانات الترخيص غير مكتملة."

        try:
            expiry_date = datetime.date.fromisoformat(expiry)
            if expiry_date < datetime.date.today():
                return False, "⛔ انتهت صلاحية الترخيص."
        except:
            return False, "⛔ تاريخ انتهاء غير صالح."

        current_device = get_device_id()
        if device_id != current_device:
            return False, "⛔ هذا الترخيص غير مخصص لهذا الجهاز."

        if not verify_license_signature():
            return False, "⛔ مفتاح الترخيص تم التلاعب به!"

        return True, "✅ مرخص (محمي)"

    elif status == "trial":
        return is_trial_valid(data)
    else:
        return False, "⛔ حالة الترخيص غير معروفة"

# ---------------------------- دوال التشفير البسيط ----------------------------
def encrypt_key(key: str) -> str:
    return base64.b64encode(key.encode()).decode()

def decrypt_key(enc_key: str) -> str:
    return base64.b64decode(enc_key.encode()).decode()

# ---------------------------- دوال مساعدة ----------------------------
def normalize_columns(df):
    return (
        df.columns
        .str.strip()
        .str.lower()
        .str.replace(" ", "_")
    )

def sanitize_dataframe(df):
    def sanitize_cell(cell):
        if isinstance(cell, str):
            cell_strip = cell.strip()
            if re.match(r"^[=+\-@]", cell_strip):
                return "'" + cell_strip
            if re.search(r"<[^>]+>", cell_strip):
                return re.sub(r"<[^>]+>", "", cell_strip)
        return cell

    for col in df.columns:
        if df[col].dtype == "object":
            df[col] = df[col].apply(sanitize_cell)
    return df

def load_file_smart(uploaded_file, max_rows=MAX_ROWS):
    if uploaded_file is None:
        raise ValueError("لم يتم رفع أي ملف.")
    file_type = uploaded_file.name.split('.')[-1].lower()
    df = None
    try:
        if file_type == 'csv':
            try:
                df = pd.read_csv(uploaded_file, sep=None, engine='python', nrows=max_rows, encoding='utf-8-sig')
            except:
                df = pd.read_csv(uploaded_file, sep=None, engine='python', nrows=max_rows, encoding='latin1')
        elif file_type in ['xlsx', 'xls']:
            engine = 'openpyxl' if file_type == 'xlsx' else 'xlrd'
            df = pd.read_excel(uploaded_file, engine=engine, nrows=max_rows)
        else:
            raise ValueError("صيغة الملف غير مدعومة. استخدم CSV أو Excel.")
    except Exception as e:
        raise ValueError(f"حدث خطأ أثناء قراءة الملف: {e}")

    if df is None or df.empty:
        raise ValueError("الملف فارغ أو لا يحتوي على بيانات.")

    if df.shape[0] > 100000:
        st.warning("⚠️ الملف كبير جداً (>100,000 صف). سيتم معالجة أول 50,000 صف فقط. يُرجى تقسيم الملف إذا لزم الأمر.")
        df = df.head(50000)

    if df.shape[1] > MAX_COLS:
        raise ValueError(f"❌ عدد الأعمدة كبير جدًا، الحد الأقصى هو {MAX_COLS} عمودًا.")

    df = df.dropna(axis=1, how='all')
    df = df.loc[:, ~df.columns.str.strip().duplicated()]
    df = sanitize_dataframe(df)
    df._normalized_columns = normalize_columns(df).tolist()
    return df, file_type

def auto_map_columns_smart(df, synonyms, standard_columns):
    mapping = {std: None for std in standard_columns.keys()}
    if not hasattr(df, '_normalized_columns'):
        normalized = normalize_columns(df).tolist()
    else:
        normalized = df._normalized_columns
    norm_to_orig = dict(zip(normalized, df.columns))

    for std_col, syn_list in synonyms.items():
        all_syn = [std_col] + syn_list
        normalized_syns = [syn.strip().lower().replace(" ", "_") for syn in all_syn]
        found = False

        for norm_syn in normalized_syns:
            if norm_syn in norm_to_orig:
                mapping[std_col] = norm_to_orig[norm_syn]
                found = True
                break
        if found:
            continue

        possible_matches = []
        for norm_syn in normalized_syns:
            if len(norm_syn) < 4:
                continue
            for norm_col in normalized:
                if norm_col.startswith(norm_syn) or norm_col.endswith(norm_syn):
                    possible_matches.append(norm_col)

        possible_matches = list(set(possible_matches))
        if len(possible_matches) == 1:
            mapping[std_col] = norm_to_orig[possible_matches[0]]

    return mapping

def validate_mapping(mapping, required):
    return [col for col in required if mapping.get(col) is None]

# ---------------------------- دوال الحساب الآمنة ----------------------------
def safe_divide(numerator, denominator):
    try:
        if pd.isna(denominator) or denominator == 0:
            return 0
        return numerator / denominator
    except Exception:
        return 0

def calculate_kpis_safe(row):
    impressions = row.get('impressions', 0)
    clicks = row.get('clicks', 0)
    spend = row.get('spend', 0)
    conversions = row.get('conversions', 0)
    revenue = row.get('revenue', 0)

    ctr = safe_divide(clicks, impressions) * 100
    cpc = safe_divide(spend, clicks)
    cpa = safe_divide(spend, conversions)
    roas = safe_divide(revenue, spend)

    return pd.Series([ctr, cpc, cpa, roas], index=['CTR', 'CPC', 'CPA', 'ROAS'])

def calculate_kpis(df, mapping):
    data = {}
    numeric_cols = ['impressions', 'clicks', 'spend', 'conversions', 'revenue']
    for std, orig in mapping.items():
        if orig and orig in df.columns:
            col_data = pd.to_numeric(df[orig], errors='coerce').fillna(0)
            if std in numeric_cols:
                col_data = col_data.clip(lower=0)
            data[std] = col_data
        else:
            if std in numeric_cols:
                data[std] = pd.Series([0] * len(df))
            else:
                data[std] = pd.Series([""] * len(df))

    if 'revenue' not in data:
        data['revenue'] = pd.Series([0] * len(df))

    df_clean = pd.DataFrame(data)
    df_clean = df_clean.dropna(subset=['campaign_name'])
    df_clean[['CTR', 'CPC', 'CPA', 'ROAS']] = df_clean.apply(calculate_kpis_safe, axis=1)

    stats = {
        'total_impressions': df_clean['impressions'].sum(),
        'total_clicks': df_clean['clicks'].sum(),
        'total_spend': df_clean['spend'].sum(),
        'total_conversions': df_clean['conversions'].sum(),
        'total_revenue': df_clean['revenue'].sum(),
        'avg_CTR': df_clean['CTR'].mean(),
        'avg_CPC': df_clean['CPC'].mean(),
        'avg_CPA': df_clean['CPA'].mean(),
        'avg_ROAS': df_clean['ROAS'].mean(),
        'num_campaigns': df_clean['campaign_name'].nunique()
    }

    best_campaign = "غير متوفر"
    worst_campaign = "غير متوفر"
    if df_clean['ROAS'].notna().any() and df_clean['ROAS'].sum() > 0:
        best_campaign = df_clean.loc[df_clean['ROAS'].idxmax(), 'campaign_name']
        worst_campaign = df_clean.loc[df_clean['ROAS'].idxmin(), 'campaign_name']
    elif df_clean['CPA'].notna().any() and df_clean['CPA'].sum() > 0:
        best_campaign = df_clean.loc[df_clean['CPA'].idxmin(), 'campaign_name']
        worst_campaign = df_clean.loc[df_clean['CPA'].idxmax(), 'campaign_name']

    stats['best_campaign'] = best_campaign
    stats['worst_campaign'] = worst_campaign
    return df_clean, stats

@st.cache_data(show_spinner=False)
def cached_calculate(df, mapping_tuple):
    mapping = dict(mapping_tuple)
    return calculate_kpis(df, mapping)

# ---------------------------- دوال التخزين المؤقت للملفات ----------------------------
def get_file_hash(uploaded_file):
    return hashlib.md5(uploaded_file.getvalue()).hexdigest()

# ---------------------------- دالة تصدير Excel المحسّنة ----------------------------
def export_excel_with_summary(df_clean, user_password=None):
    from openpyxl.styles import Font, PatternFill, Alignment, Protection
    from openpyxl.utils import get_column_letter
    from openpyxl.formatting.rule import CellIsRule
    from openpyxl.worksheet.table import Table, TableStyleInfo

    output = BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df_clean.to_excel(writer, sheet_name='تفصيلي', index=False)
        wb = writer.book
        ws_detail = wb['تفصيلي']

        col_letter = {}
        for col_idx, col_name in enumerate(df_clean.columns, start=1):
            col_letter[col_name] = get_column_letter(col_idx)

        required_cols = ['impressions', 'clicks', 'spend', 'conversions', 'revenue', 'CTR', 'CPC', 'CPA', 'ROAS']
        for col in required_cols:
            if col not in col_letter:
                raise ValueError(f"العمود {col} غير موجود في البيانات المصدرة")

        for row in range(2, ws_detail.max_row + 1):
            ws_detail[f"{col_letter['CTR']}{row}"].value = f"=IFERROR({col_letter['clicks']}{row}/{col_letter['impressions']}{row},0)"
            ws_detail[f"{col_letter['CPC']}{row}"].value = f"=IFERROR({col_letter['spend']}{row}/{col_letter['clicks']}{row},0)"
            ws_detail[f"{col_letter['CPA']}{row}"].value = f"=IFERROR({col_letter['spend']}{row}/{col_letter['conversions']}{row},0)"
            ws_detail[f"{col_letter['ROAS']}{row}"].value = f"=IFERROR({col_letter['revenue']}{row}/{col_letter['spend']}{row},0)"

        header_fill_detail = PatternFill(start_color='FFC000', end_color='FFC000', fill_type='solid')
        for col in range(1, ws_detail.max_column + 1):
            cell = ws_detail[f"{get_column_letter(col)}1"]
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = header_fill_detail
            cell.alignment = Alignment(horizontal='center')

        for row in range(2, ws_detail.max_row + 1):
            ws_detail[f"{col_letter['impressions']}{row}"].number_format = '#,##0'
            ws_detail[f"{col_letter['clicks']}{row}"].number_format = '#,##0'
            ws_detail[f"{col_letter['spend']}{row}"].number_format = '$#,##0.00'
            ws_detail[f"{col_letter['conversions']}{row}"].number_format = '#,##0'
            ws_detail[f"{col_letter['revenue']}{row}"].number_format = '$#,##0.00'
            ws_detail[f"{col_letter['CTR']}{row}"].number_format = '0.00%'
            ws_detail[f"{col_letter['CPC']}{row}"].number_format = '0.00'
            ws_detail[f"{col_letter['CPA']}{row}"].number_format = '0.00'
            ws_detail[f"{col_letter['ROAS']}{row}"].number_format = '0.00'

        if ws_detail.max_row > 1:
            green_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
            red_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')

            roas_col = col_letter['ROAS']
            roas_range = f"{roas_col}2:{roas_col}{ws_detail.max_row}"
            ws_detail.conditional_formatting.add(roas_range, CellIsRule(operator='greaterThan', formula=['1'], fill=green_fill))
            ws_detail.conditional_formatting.add(roas_range, CellIsRule(operator='lessThan', formula=['1'], fill=red_fill))

            ctr_col = col_letter['CTR']
            ctr_range = f"{ctr_col}2:{ctr_col}{ws_detail.max_row}"
            ws_detail.conditional_formatting.add(ctr_range, CellIsRule(operator='greaterThan', formula=['0.05'], fill=green_fill))
            ws_detail.conditional_formatting.add(ctr_range, CellIsRule(operator='lessThan', formula=['0.01'], fill=red_fill))

        if ws_detail.max_row > 1:
            first_col = get_column_letter(1)
            last_col = col_letter[df_clean.columns[-1]]
            table_ref = f"{first_col}1:{last_col}{ws_detail.max_row}"
            ad_table = Table(displayName="AdInsightTable", ref=table_ref)
            style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False, showLastColumn=False, showRowStripes=True, showColumnStripes=False)
            ad_table.tableStyleInfo = style
            ws_detail.add_table(ad_table)
            ws_detail.auto_filter.ref = table_ref
            ws_detail.freeze_panes = f"{first_col}2"

        for row in ws_detail.iter_rows():
            for cell in row:
                cell.protection = Protection(locked=True)

        raw_columns = ['campaign_name', 'impressions', 'clicks', 'spend', 'conversions', 'revenue']
        for col_name in raw_columns:
            if col_name in col_letter:
                col_let = col_letter[col_name]
                for row in range(2, ws_detail.max_row + 1):
                    ws_detail[f"{col_let}{row}"].protection = Protection(locked=False)

        ws_detail.protection.formatCells = True
        ws_detail.protection.enable()

        summary_data = {
            'المقياس': ['إجمالي مرات الظهور', 'إجمالي النقرات', 'إجمالي الإنفاق', 'إجمالي التحويلات', 'إجمالي الإيرادات', 'متوسط CTR', 'متوسط CPC', 'متوسط CPA', 'متوسط ROAS'],
            'القيمة': [
                f"=SUM({col_letter['impressions']}2:{col_letter['impressions']}{ws_detail.max_row})",
                f"=SUM({col_letter['clicks']}2:{col_letter['clicks']}{ws_detail.max_row})",
                f"=SUM({col_letter['spend']}2:{col_letter['spend']}{ws_detail.max_row})",
                f"=SUM({col_letter['conversions']}2:{col_letter['conversions']}{ws_detail.max_row})",
                f"=SUM({col_letter['revenue']}2:{col_letter['revenue']}{ws_detail.max_row})",
                f"=IFERROR(SUM({col_letter['CTR']}2:{col_letter['CTR']}{ws_detail.max_row})/COUNTA({col_letter['CTR']}2:{col_letter['CTR']}{ws_detail.max_row}),0)",
                f"=IFERROR(SUM({col_letter['CPC']}2:{col_letter['CPC']}{ws_detail.max_row})/COUNTA({col_letter['CPC']}2:{col_letter['CPC']}{ws_detail.max_row}),0)",
                f"=IFERROR(SUM({col_letter['CPA']}2:{col_letter['CPA']}{ws_detail.max_row})/COUNTA({col_letter['CPA']}2:{col_letter['CPA']}{ws_detail.max_row}),0)",
                f"=IFERROR(SUM({col_letter['ROAS']}2:{col_letter['ROAS']}{ws_detail.max_row})/COUNTA({col_letter['ROAS']}2:{col_letter['ROAS']}{ws_detail.max_row}),0)"
            ]
        }
        df_summary = pd.DataFrame(summary_data)
        df_summary.to_excel(writer, sheet_name='Summary', index=False)
        ws_summary = wb['Summary']

        header_fill_summary = PatternFill(start_color='00B0F0', end_color='00B0F0', fill_type='solid')
        for col in range(1, ws_summary.max_column + 1):
            cell = ws_summary[f"{get_column_letter(col)}1"]
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = header_fill_summary
            cell.alignment = Alignment(horizontal='center')

        for row in range(2, ws_summary.max_row + 1):
            metric_name = str(ws_summary[f'A{row}'].value or "")
            if "CTR" in metric_name:
                ws_summary[f'B{row}'].number_format = '0.00%'
            elif "الإنفاق" in metric_name or "الإيرادات" in metric_name:
                ws_summary[f'B{row}'].number_format = '$#,##0.00'
            elif "CPC" in metric_name or "CPA" in metric_name or "ROAS" in metric_name:
                ws_summary[f'B{row}'].number_format = '0.00'
            else:
                ws_summary[f'B{row}'].number_format = '#,##0'

        for row in ws_summary.iter_rows():
            for cell in row:
                cell.protection = Protection(locked=True)

        ws_summary.protection.enable()

        if ws_detail.max_row > 1:
            if user_password:
                password = user_password
            else:
                password = uuid.uuid4().hex[:8]
            ws_detail.protection.set_password(password)
            ws_summary.protection.set_password(password)
            if not user_password:
                st.info(f"🔐 تم حماية ملف Excel بكلمة مرور عشوائية: `{password}`\nاحتفظ بها لأنها لن تظهر مرة أخرى.")
            else:
                st.info("🔐 تم حماية ملف Excel بكلمة المرور التي أدخلتها.")

        for col in ws_summary.columns:
            max_length = max(len(str(cell.value)) for cell in col)
            adjusted_width = min(max_length + 2, 30)
            ws_summary.column_dimensions[get_column_letter(col[0].column)].width = adjusted_width

    output.seek(0)
    return output

# ---------------------------- دالة توليد الملخص الذكي ----------------------------
def generate_ai_summary_safe(stats, model="gpt-3.5-turbo", max_retries=3):
    api_key = st.session_state.get('api_key')
    if not api_key:
        return None, None, "❌ مفتاح API غير موجود."

    client = OpenAI(api_key=api_key, timeout=30)

    prompt = f"""
أنت خبير تسويق رقمي. بناءً على إحصائيات الحملات التالية، أكتب:
1- ملخص تنفيذي موجز (3-4 جمل) باللغة العربية.
2- ثلاث توصيات عملية قابلة للتنفيذ (على شكل قائمة نقاط).

الإحصائيات:
- إجمالي مرات الظهور: {stats['total_impressions']:,.0f}
- إجمالي النقرات: {stats['total_clicks']:,.0f}
- إجمالي الإنفاق: ${stats['total_spend']:,.2f}
- إجمالي التحويلات: {stats['total_conversions']:,.0f}
- إجمالي الإيرادات: ${stats['total_revenue']:,.2f}
- متوسط CTR: {stats['avg_CTR']:.2f}%
- متوسط CPC: ${stats['avg_CPC']:.2f}
- متوسط CPA: ${stats['avg_CPA']:.2f}
- متوسط ROAS: {stats['avg_ROAS']:.2f}
- عدد الحملات: {stats['num_campaigns']}
- أفضل حملة: {stats['best_campaign']}
- أسوأ حملة: {stats['worst_campaign']}

فصل الملخص عن التوصيات باستخدام ===
"""

    attempt = 0
    while attempt < max_retries:
        try:
            response = client.chat.completions.create(
                model=model,
                messages=[
                    {"role": "system", "content": "أنت مساعد خبير في التسويق الرقمي."},
                    {"role": "user", "content": prompt}
                ],
                temperature=0.7,
                max_tokens=600,
            )

            if not response.choices:
                raise ValueError("❌ استجابة OpenAI فارغة أو غير متوقعة.")
            content = response.choices[0].message.content.strip()
            if not content:
                raise ValueError("❌ محتوى الاستجابة فارغ.")

            if "===" in content:
                parts = content.split("===")
                summary_text = parts[0].strip()
                recommendations_text = parts[1].strip() if len(parts) > 1 else ""
            else:
                summary_text = content
                recommendations_text = ""

            return summary_text, recommendations_text, None

        except AuthenticationError:
            return None, None, "❌ فشل المصادقة: تحقق من مفتاح API."
        except RateLimitError:
            attempt += 1
            if attempt >= max_retries:
                return None, None, "⚠️ تم تجاوز الحد المسموح للطلبات."
            sleep_time = (2 ** attempt) + random.uniform(0, 1)
            time.sleep(sleep_time)
        except APIConnectionError:
            attempt += 1
            if attempt >= max_retries:
                return None, None, "⚠️ فشل الاتصال بـ OpenAI بعد عدة محاولات."
            sleep_time = (2 ** attempt) + random.uniform(0, 1)
            time.sleep(sleep_time)
        except APIError:
            return None, None, "⚠️ خطأ في خوادم OpenAI. حاول مرة أخرى لاحقاً."
        except Exception as e:
            return None, None, f"⚠️ خطأ غير متوقع: {e}"

    return None, None, "⚠️ لم يتم توليد ملخص ذكي. سيتم استخدام الملخص الافتراضي."

def generate_default_summary(stats):
    return f"""
**ملخص تنفيذي (افتراضي)**

- إجمالي مرات الظهور: {stats['total_impressions']:,.0f}
- إجمالي النقرات: {stats['total_clicks']:,.0f}
- إجمالي الإنفاق: ${stats['total_spend']:,.2f}
- إجمالي التحويلات: {stats['total_conversions']:,.0f}
- إجمالي الإيرادات: ${stats['total_revenue']:,.2f}

متوسط نسبة النقر إلى الظهور (CTR): {stats['avg_CTR']:.2f}%
متوسط تكلفة النقرة (CPC): ${stats['avg_CPC']:.2f}
متوسط تكلفة التحويل (CPA): ${stats['avg_CPA']:.2f}
متوسط العائد على الإنفاق (ROAS): {stats['avg_ROAS']:.2f}

أفضل حملة: {stats['best_campaign']}
أسوأ حملة: {stats['worst_campaign']}
"""

def generate_pdf_report(df, stats, summary_text, recommendations_text, logo_path=None):
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, rightMargin=72, leftMargin=72, topMargin=72, bottomMargin=18)
    elements = []

    styles = getSampleStyleSheet()
    style_normal = styles['Normal']
    style_heading = styles['Heading1']
    style_sub = styles['Heading2']

    title = Paragraph("تقرير أداء الحملات - AdInsight AI", style_heading)
    elements.append(title)
    elements.append(Spacer(1, 12))

    date_str = datetime.datetime.now().strftime("%Y-%m-%d %H:%M")
    date_para = Paragraph(f"تاريخ الإنشاء: {date_str}", style_normal)
    elements.append(date_para)
    elements.append(Spacer(1, 24))

    if logo_path:
        try:
            img = ImageReader(logo_path)
            iw, ih = img.getSize()
            max_width_cm = 6 * cm
            max_height_cm = 3 * cm
            scale = min(max_width_cm / iw, max_height_cm / ih, 1.0)
            logo_width = iw * scale
            logo_height = ih * scale
            logo = Image(logo_path, width=logo_width, height=logo_height)
            logo.hAlign = 'RIGHT'
            elements.append(logo)
            elements.append(Spacer(1, 12))
        except Exception as e:
            print(f"تحذير: فشل إضافة الشعار - {e}")

    elements.append(Paragraph("ملخص تنفيذي", style_sub))
    elements.append(Spacer(1, 6))

    if summary_text:
        for line in summary_text.split('\n'):
            if line.strip():
                p = Paragraph(line.strip(), style_normal)
                elements.append(p)
    else:
        default_summary = generate_default_summary(stats)
        for line in default_summary.split('\n'):
            if line.strip():
                p = Paragraph(line.strip(), style_normal)
                elements.append(p)

    elements.append(Spacer(1, 24))

    elements.append(Paragraph("مؤشرات الأداء الرئيسية", style_sub))
    elements.append(Spacer(1, 6))

    table_data = [
        ['المقياس', 'القيمة'],
        ['إجمالي مرات الظهور', f"{stats['total_impressions']:,.0f}"],
        ['إجمالي النقرات', f"{stats['total_clicks']:,.0f}"],
        ['إجمالي الإنفاق', f"${stats['total_spend']:,.2f}"],
        ['إجمالي التحويلات', f"{stats['total_conversions']:,.0f}"],
        ['إجمالي الإيرادات', f"${stats['total_revenue']:,.2f}"],
        ['متوسط CTR', f"{stats['avg_CTR']:.2f}%"],
        ['متوسط CPC', f"${stats['avg_CPC']:.2f}"],
        ['متوسط CPA', f"${stats['avg_CPA']:.2f}"],
        ['متوسط ROAS', f"{stats['avg_ROAS']:.2f}"],
    ]

    table = Table(table_data, colWidths=[150, 150])
    table.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (1,0), colors.grey),
        ('TEXTCOLOR', (0,0), (1,0), colors.whitesmoke),
        ('ALIGN', (0,0), (-1,-1), 'CENTER'),
        ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
        ('FONTSIZE', (0,0), (-1,0), 12),
        ('BOTTOMPADDING', (0,0), (-1,0), 8),
        ('BACKGROUND', (0,1), (-1,-1), colors.beige),
        ('GRID', (0,0), (-1,-1), 1, colors.black)
    ]))
    elements.append(table)
    elements.append(Spacer(1, 24))

    # جدول تفصيلي ديناميكي
    elements.append(Paragraph("تفاصيل الحملات (أول 10)", style_sub))
    elements.append(Spacer(1, 6))

    required_cols = ['campaign_name', 'impressions', 'clicks', 'spend', 'conversions']
    optional_cols = ['revenue', 'CTR', 'CPC', 'CPA', 'ROAS']
    available_cols = df.columns.tolist()
    display_cols = [col for col in required_cols if col in available_cols] + \
                   [col for col in optional_cols if col in available_cols]

    if not display_cols:
        display_cols = ['campaign_name'] if 'campaign_name' in available_cols else available_cols[:1]

    display_df = df[display_cols].head(10).copy()

    for col in display_df.columns:
        if col in ['CTR', 'CPC', 'CPA', 'ROAS']:
            display_df[col] = display_df[col].apply(lambda x: f"{x:.2f}" if pd.notna(x) else "N/A")
        elif col in ['spend', 'revenue']:
            display_df[col] = display_df[col].apply(lambda x: f"${x:,.2f}" if pd.notna(x) else "$0.00")
        elif col in ['impressions', 'clicks', 'conversions']:
            display_df[col] = display_df[col].apply(lambda x: f"{x:,.0f}" if pd.notna(x) else "0")
        else:
            display_df[col] = display_df[col].astype(str)

    header = list(display_df.columns)
    data_rows = display_df.values.tolist()
    table_data = [header] + data_rows

    table2 = Table(table_data, repeatRows=1)
    table2.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), colors.grey),
        ('TEXTCOLOR', (0,0), (-1,0), colors.whitesmoke),
        ('ALIGN', (0,0), (-1,-1), 'CENTER'),
        ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
        ('FONTSIZE', (0,0), (-1,0), 10),
        ('BOTTOMPADDING', (0,0), (-1,0), 6),
        ('BACKGROUND', (0,1), (-1,-1), colors.white),
        ('GRID', (0,0), (-1,-1), 1, colors.black)
    ]))
    elements.append(table2)
    elements.append(Spacer(1, 12))

    if recommendations_text:
        elements.append(Paragraph("توصيات ذكية", style_sub))
        elements.append(Spacer(1, 6))
        for line in recommendations_text.split('\n'):
            if line.strip():
                p = Paragraph(line.strip(), style_normal)
                elements.append(p)
    else:
        elements.append(Paragraph("توصيات ذكية", style_sub))
        elements.append(Spacer(1, 6))
        rec_text = """
        - زيادة ميزانية الحملات ذات ROAS المرتفع.
        - تحسين الجمهور المستهدف للحملات ذات CPA المرتفع.
        - مراجعة تصميم الإعلانات للحملات ذات نسبة النقر المنخفضة.
        """
        for line in rec_text.strip().split('\n'):
            if line.strip():
                p = Paragraph(line.strip(), style_normal)
                elements.append(p)

    doc.build(elements)
    buffer.seek(0)
    return buffer

# ---------------------------- دالة التحقق من imports عبر AST ----------------------------
def validate_no_imports(code: str):
    tree = ast.parse(code)
    for node in ast.walk(tree):
        if isinstance(node, (ast.Import, ast.ImportFrom)):
            raise ValueError("Import statements are not allowed.")

# ---------------------------- منفذ الأكواد الآمن ----------------------------
class CodeExecutor:
    def __init__(self):
        safe_builtins = {
            "len": len,
            "sum": sum,
            "min": min,
            "max": max,
            "range": range,
            "abs": abs,
            "round": round,
            "sorted": sorted,
            "list": list,
            "dict": dict,
            "enumerate": enumerate,
        }
        self._environment = {
            "__builtins__": safe_builtins,
        }
        self._environment.update({
            "pd": pd,
            "np": np,
            "df": None,
        })

    def execute(self, code: str, data_frame):
        self._environment["df"] = data_frame

        try:
            validate_no_imports(code)
        except Exception as e:
            return {"error": f"Validation error: {str(e)}"}

        try:
            exec(code, self._environment)
            result = self._environment.get("result", None)
            return result
        except Exception as e:
            return {"error": f"Execution error: {str(e)}"}

executor = CodeExecutor()

# ---------------------------- دوال واجهة الترخيص ----------------------------
def render_license_status(data):
    st.sidebar.header("🔑 حالة الترخيص")
    status = data.get("license_status", "trial")
    usage = data.get("usage_count", 0)
    limit = data.get("trial_limit", 10)
    first_use = data.get("first_use")
    trial_days = data.get("trial_days", 7)

    if status == "active":
        expiry = data.get("expiry", "غير محدد")
        st.sidebar.success(f"✅ مرخص\nتاريخ الانتهاء: {expiry}")
    elif status == "trial":
        from datetime import datetime
        try:
            first_use_date = datetime.fromisoformat(first_use)
            delta_days = (datetime.now() - first_use_date).days
            remaining_days = max(trial_days - delta_days, 0)
        except:
            remaining_days = trial_days
        remaining_usage = max(limit - usage, 0)
        st.sidebar.warning(f"⚠️ نسخة تجريبية:\n- {remaining_usage} استخدامات متبقية\n- {remaining_days} أيام متبقية")
    else:
        st.sidebar.error("⛔ حالة الترخيص غير صالحة")

def render_license_activation():
    with st.sidebar.expander("🔓 تفعيل الترخيص"):
        key_input = st.text_input("أدخل مفتاح الترخيص", type="password")
        if st.button("تفعيل"):
            success, msg = activate_license(key_input)
            if success:
                st.success(msg)
                st.experimental_rerun()
            else:
                st.error(msg)

def get_trial_notifications(data):
    usage = data.get("usage_count", 0)
    limit = data.get("trial_limit", 10)
    first_use_str = data.get("first_use")
    trial_days = data.get("trial_days", 7)

    from datetime import datetime
    try:
        first_use_date = datetime.fromisoformat(first_use_str)
        delta_days = (datetime.now() - first_use_date).days
    except:
        delta_days = 0

    remaining_usage = max(limit - usage, 0)
    remaining_days = max(trial_days - delta_days, 0)

    messages = []
    if remaining_days <= 2:
        messages.append(f"⚠️ النسخة التجريبية ستنتهي بعد {remaining_days} أيام!")
    if remaining_usage <= 2:
        messages.append(f"⚠️ تبقى {remaining_usage} استخدامات فقط!")

    return messages

def render_trial_notifications():
    try:
        data = load_client_data_encrypted()
    except:
        return
    if data.get("license_status") == "trial":
        notifications = get_trial_notifications(data)
        for msg in notifications:
            st.sidebar.warning(msg)

# ---------------------------- تهيئة الجلسة (محدثة) ----------------------------
def init_session():
    defaults = {
        "df": None,
        "last_file_hash": None,
        "mapping": None,
        "run_analysis": False,
        "stats": None,
        "df_clean": None,
        "ai_summary": None,
        "ai_recs": None,
        "ai_error": None,
        "pdf_buffer": None
    }
    for key, value in defaults.items():
        if key not in st.session_state:
            st.session_state[key] = value

# ---------------------------- قسم رفع الملف الذكي (محدث) ----------------------------
def upload_section():
    uploaded_file = st.file_uploader("ارفع ملف CSV أو Excel", type=['csv', 'xlsx', 'xls'])

    if uploaded_file:
        current_hash = get_file_hash(uploaded_file)

        if st.session_state.last_file_hash != current_hash:
            try:
                df, file_type = load_file_smart(uploaded_file)

                st.session_state.df = df
                st.session_state.last_file_hash = current_hash
                st.session_state.mapping = auto_map_columns_smart(df, SYNONYMS, STANDARD_COLUMNS)

                # إعادة ضبط حالة التحليل والنتائج السابقة لأن البيانات تغيرت
                st.session_state.run_analysis = False
                st.session_state.stats = None
                st.session_state.df_clean = None
                st.session_state.ai_summary = None
                st.session_state.ai_recs = None
                st.session_state.ai_error = None
                st.session_state.pdf_buffer = None

                st.success("✅ تم التعرف على ملف جديد ومعالجته بنجاح.")
            except Exception as e:
                st.error(f"❌ حدث خطأ أثناء المعالجة: {e}")
                st.stop()

# ---------------------------- واجهة Streamlit ----------------------------
def main():
    st.set_page_config(page_title="AdInsight AI", layout="wide")

    # تهيئة الجلسة
    init_session()

    st.title("📊 AdInsight AI - مولد تقارير الحملات الإعلانية مع الذكاء الاصطناعي")
    st.markdown("---")

    # -------------------- التحقق من الترخيص --------------------
    is_allowed, message = check_license_secure_with_trial()

    if not is_allowed:
        st.error(f"⛔ {message}")
        render_license_activation()
        st.stop()
    else:
        if DEV_MODE:
            st.info("🚧 وضع التطوير: الترخيص غير مفعل.")
        else:
            st.info(message)

    # -------------------- عرض حالة الترخيص --------------------
    try:
        data = load_client_data_encrypted()
        render_license_status(data)
    except:
        if DEV_MODE:
            st.sidebar.info("🔓 وضع التطوير: الترخيش غير مفعل.")
    render_trial_notifications()

    # -------------------- إعدادات الشريط الجانبي --------------------
    if 'api_key_encrypted' not in st.session_state:
        st.session_state['api_key_encrypted'] = None
    if 'logo_path' not in st.session_state:
        st.session_state['logo_path'] = None

    with st.sidebar:
        st.header("الإعدادات")

        logo_file = st.file_uploader("شعار العميل (اختياري - White Label)", type=['png', 'jpg', 'jpeg'])
        if logo_file is not None:
            if st.session_state['logo_path'] and os.path.exists(st.session_state['logo_path']):
                os.remove(st.session_state['logo_path'])
            logo_filename = os.path.join(tempfile.gettempdir(), f"temp_logo_{uuid.uuid4().hex}.png")
            with open(logo_filename, "wb") as f:
                f.write(logo_file.getbuffer())
            st.session_state['logo_path'] = logo_filename
        else:
            st.session_state['logo_path'] = None

        st.markdown("---")
        st.subheader("🔑 مفتاح OpenAI API")
        api_key_input = st.text_input(
            "أدخل مفتاح API الخاص بك (للملخصات المتقدمة)",
            type="password",
            help="يمكنك الحصول على مفتاح من platform.openai.com."
        )
        if api_key_input:
            st.session_state['api_key_encrypted'] = encrypt_key(api_key_input)
        else:
            st.session_state['api_key_encrypted'] = None

        if st.button("🔒 حذف المفتاح من الجلسة"):
            st.session_state['api_key_encrypted'] = None
            st.success("✅ تم حذف المفتاح.")

        st.markdown("---")
        st.subheader("💰 نموذج التسعير (محاكاة)")
        plan = st.radio("اختر الباقة", ["Basic (29$)", "Pro (59$)", "Agency (99$)"])
        st.info("✅ الباقة Pro و Agency تدعمان الملخصات المتقدمة بالذكاء الاصطناعي.")

        if plan == "Basic (29$)":
            model_option = None
            st.caption("الباقة الأساسية لا تدعم الذكاء الاصطناعي.")
        elif plan == "Pro (59$)":
            model_option = "gpt-3.5-turbo"
            st.caption("باقة Pro تستخدم نموذج GPT-3.5-Turbo.")
        elif plan == "Agency (99$)":
            model_option = st.selectbox("نموذج الذكاء الاصطناعي", ["gpt-3.5-turbo", "gpt-4"], index=0)

        st.markdown("---")
        excel_password = st.text_input("كلمة مرور Excel (اختياري)", type="password",
                                       help="اتركه فارغاً لإنشاء كلمة عشوائية.")
        st.markdown("---")
        st.caption("AdInsight AI - جميع الحقوق محفوظة © 2026")

    # -------------------- رفع الملف --------------------
    upload_section()

    # استرجاع البيانات من الجلسة
    df = st.session_state.df
    mapping = st.session_state.mapping

    if df is None:
        st.info("📁 يرجى رفع ملف للبدء.")
        st.stop()

    # -------------------- بعد التحميل --------------------
    st.subheader("🔍 عينة من البيانات المرفوعة")
    st.dataframe(df.head())

    # التحقق من الأعمدة المفقودة
    missing = validate_mapping(mapping, REQUIRED_COLUMNS)

    if missing:
        st.warning(f"⚠️ لم نتمكن من تحديد الأعمدة التالية تلقائياً: {', '.join(missing)}")
        st.markdown("يرجى تحديد العمود المناسب لكل منها من القوائم أدناه.")
        manual_mapping = {}
        for std_col in missing:
            options = ['-- لا يوجد --'] + list(df.columns)
            selected = st.selectbox(
                label=f"**{STANDARD_COLUMNS[std_col]}**",
                options=options,
                help=f"ابحث عن عمود يحتوي على بيانات {STANDARD_COLUMNS[std_col]}، مثل: {', '.join(SYNONYMS[std_col][:3])} ...",
                key=f"manual_{std_col}"
            )
            if selected != '-- لا يوجد --':
                manual_mapping[std_col] = selected
        for std, orig in manual_mapping.items():
            mapping[std] = orig
        missing_after = validate_mapping(mapping, REQUIRED_COLUMNS)
        if missing_after:
            st.error("❌ لا يزال هناك أعمدة إلزامية مفقودة. يرجى التحقق من التعيينات أعلاه.")
            st.stop()
        else:
            st.success("✅ تم تعيين جميع الأعمدة بنجاح!")

    st.subheader("📌 التعيينات النهائية")
    mapping_df = pd.DataFrame([
        {"العمود القياسي": STANDARD_COLUMNS[std], "العمود في الملف": orig if orig else "❌ مفقود"}
        for std, orig in mapping.items()
    ])
    st.table(mapping_df)

    # -------------------- زر التشغيل الرئيسي --------------------
    if st.button("🚀 تحليل البيانات وإنشاء التقرير"):
        st.session_state.run_analysis = True
        st.session_state.stats = None          # تصفير لإجبار إعادة الحساب
        st.session_state.df_clean = None
        st.session_state.ai_summary = None
        st.session_state.ai_recs = None
        st.session_state.ai_error = None
        st.session_state.pdf_buffer = None     # تصفير لإعادة توليد التقرير الجديد

    # -------------------- عرض النتائج --------------------
    if st.session_state.run_analysis:

        # 1. تنفيذ الحسابات مرة واحدة فقط
        if st.session_state.stats is None:
            with st.spinner("جاري تحليل البيانات..."):
                df_clean, stats = cached_calculate(df, tuple(mapping.items()))
                st.session_state.df_clean = df_clean
                st.session_state.stats = stats

            # 2. تنفيذ استدعاء الذكاء الاصطناعي مرة واحدة فقط
            if model_option is not None:
                if st.session_state.get('api_key_encrypted'):
                    with st.spinner("🧠 جاري توليد الملخص بالذكاء الاصطناعي..."):
                        summ, recs, err = generate_ai_summary_safe(stats, model=model_option)
                        st.session_state.ai_summary = summ
                        st.session_state.ai_recs = recs
                        st.session_state.ai_error = err
                else:
                    st.session_state.ai_error = "❌ مفتاح API غير موجود. يرجى إدخاله في الشريط الجانبي."
            else:
                st.session_state.ai_error = "ℹ️ الباقة الحالية لا تدعم الذكاء الاصطناعي. استخدم الملخص الافتراضي."

        df_clean = st.session_state.df_clean
        stats = st.session_state.stats

        st.subheader("📈 لوحة التحليل")
        col1, col2, col3, col4 = st.columns(4)
        col1.metric("إجمالي مرات الظهور", f"{stats['total_impressions']:,.0f}")
        col2.metric("إجمالي النقرات", f"{stats['total_clicks']:,.0f}")
        col3.metric("إجمالي الإنفاق", f"${stats['total_spend']:,.2f}")
        col4.metric("إجمالي التحويلات", f"{stats['total_conversions']:,.0f}")

        col1, col2, col3, col4 = st.columns(4)
        col1.metric("متوسط CTR", f"{stats['avg_CTR']:.2f}%")
        col2.metric("متوسط CPC", f"${stats['avg_CPC']:.2f}")
        col3.metric("متوسط CPA", f"${stats['avg_CPA']:.2f}")
        col4.metric("متوسط ROAS", f"{stats['avg_ROAS']:.2f}")

        st.markdown(f"**🏆 أفضل حملة:** {stats['best_campaign']}")
        st.markdown(f"**📉 أسوأ حملة:** {stats['worst_campaign']}")

        st.subheader("🧠 الملخص التنفيذي")
        if st.session_state.ai_summary:
            st.markdown(st.session_state.ai_summary)
        else:
            st.markdown(generate_default_summary(stats))
            if st.session_state.ai_error:
                st.warning(st.session_state.ai_error)

        if st.session_state.ai_recs:
            st.subheader("📌 التوصيات العملية")
            st.markdown(st.session_state.ai_recs)

        st.markdown("---")
        st.subheader("📥 تحميل التقارير")
        col_pdf, col_excel = st.columns(2)

        # إصلاح مشكلة زر الـ PDF
        with col_pdf:
            if st.session_state.pdf_buffer is None:
                if st.button("📝 تجهيز تقرير PDF"):
                    with st.spinner("جاري التجهيز..."):
                        st.session_state.pdf_buffer = generate_pdf_report(
                            df_clean, stats,
                            st.session_state.ai_summary,
                            st.session_state.ai_recs,
                            st.session_state.get('logo_path')
                        )
                        st.rerun()   # لإظهار زر التحميل

            if st.session_state.pdf_buffer:
                st.download_button(
                    label="📄 تحميل التقرير (PDF)",
                    data=st.session_state.pdf_buffer,
                    file_name=f"AdInsight_Report_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf",
                    mime="application/pdf"
                )

        # زر تحميل Excel (يعمل بسلاسة لأنه لا يتطلب استدعاء خارجي ثقيل)
        with col_excel:
            excel_output = export_excel_with_summary(df_clean, user_password=excel_password if excel_password else None)
            st.download_button(
                label="📊 تحميل تقرير Excel",
                data=excel_output,
                file_name=f"AdInsight_Analysis_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )

    # -------------------- قسم تحليل مخصص (كود بايثون آمن) --------------------
    with st.expander("🧪 تحليل مخصص (كود بايثون آمن)"):
        st.markdown("اكتب كود بايثون لمعالجة DataFrame الحالي (`df`). يجب أن يخزن النتيجة في متغير `result`.")
        code_input = st.text_area("الكود", height=200, value="# مثال: result = df.head(10)")
        if st.button("تشغيل الكود"):
            try:
                if st.session_state.df_clean is not None:
                    df_to_use = st.session_state.df_clean
                else:
                    df_to_use = df
                res = executor.execute(code_input, df_to_use)
                if isinstance(res, dict) and "error" in res:
                    st.error(f"خطأ: {res['error']}")
                else:
                    st.success("تم التنفيذ بنجاح.")
                    st.write("النتيجة:", res)
            except Exception as e:
                st.error(f"خطأ غير متوقع: {e}")

if __name__ == "__main__":
    main()
